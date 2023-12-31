import datetime
import os
import sqlite3

import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ParseMode
from loguru import logger
from openpyxl.utils import get_column_letter

from database.db_manager import recording_user_data, check_user_data_exists, database, record_user_to_db
from keyboards.user_keyboards import welcome_keyboard, contact_keyboard
from messages.user_messages import message_text
from system.dispatcher import dp, bot

# Настройка логирования
logger.add("setting/log/log.log", rotation="1 MB", compression="zip")


class MyStates(StatesGroup):
    waiting_for_message = State()


@dp.message_handler(Command('send_message'))
async def send_message(message: types.Message):
    """Запрашивает текст сообщения у администратора"""
    await message.answer("Введите текст для рассылки:")
    await MyStates.waiting_for_message.set()  # Устанавливаем состояние "ожидание сообщения"


@dp.message_handler(state=MyStates.waiting_for_message, content_types=types.ContentType.TEXT)
async def process_send_message(message: types.Message, state: FSMContext):
    """
    Этот хендлер будет ждать введенного текста и выполнять рассылку
    """
    async with state.proxy() as data:
        data['message_text'] = message.text
    # Получаем список уникальных ID пользователей из базы данных
    user_ids = get_user_ids()
    if user_ids:
        # Рассылка сообщения всем пользователям из списка
        for user_id in user_ids:
            try:
                await bot.send_message(user_id, data['message_text'], parse_mode=ParseMode.MARKDOWN)
            except Exception as e:
                print(f"Ошибка при отправке сообщения пользователю {user_id}: {str(e)}")
    await message.answer("Сообщение успешно разослано всем пользователям.")
    await state.finish()


# Функция для получения уникальных ID пользователей из базы данных
def get_user_ids():
    try:
        conn = sqlite3.connect(database)  # Замените 'your_database.db' на имя вашей базы данных
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT user_id FROM user_run_bot")
        user_ids = [row[0] for row in cursor.fetchall()]
        return user_ids
    except Exception as e:
        print(f"Ошибка при получении ID пользователей из базы данных: {str(e)}")
        return []


@dp.message_handler(commands=['start'])
async def start_command(message: types.Message):
    """
    Обработчик команды /start, который отправляет приветственное сообщение и клавиатуру пользователю.
    Параметры:
    - message (types. Message): Объект сообщения пользователя, который вызвал команду.
    Внутренние действия:
    1. Создается инлайн-клавиатура с помощью функции welcome_keyboard().
    2. Если пользователь не существует в базе данных, он будет записан туда.
    3. Отправляется пользователю приветственное сообщение и клавиатура.
    Пример использования:
    После отправки команды /start, пользователь получит приветственное сообщение и клавиатуру.
    """
    try:
        user_id = message.from_user.id
        username = message.from_user.username
        first_name = message.from_user.first_name
        last_name = message.from_user.last_name
        # Если пользователя нет в базе данных, записываем его
        date_now = datetime.datetime.now()
        record_user_to_db(user_id, username, first_name, last_name, date_now)
        inline_keyboard = welcome_keyboard()
        await message.reply(message_text, reply_markup=inline_keyboard)
    except Exception as e:
        logger.exception(e)


@dp.message_handler(content_types=types.ContentType.CONTACT, state="*")
async def get_contact(message: types.Message):
    """
    Обработчик для получения контакта пользователя.
    Параметры:
    - message (types. Message): Объект сообщения, содержащий контакт пользователя.
    Внутренние действия:
    1. Извлекает информацию о контакте пользователя, такую как номер телефона и полное имя.
    2. Получает информацию о пользователе, такую как user_id, username, first_name и last_name.
    3. Проверяет наличие данных о пользователе в базе данных на основе user_id.
    4. Если данные о пользователе уже существуют, отправляет сообщение, что пользователь уже зарегистрирован.
    5. Если данные о пользователе отсутствуют, запрашивает номер телефона, записывает данные и отправляет видео.
    Пример использования:
    При отправке контакта пользователь будет зарегистрирован и получит видео, если он ещё не зарегистрирован.
    """
    contact = message.contact
    # Получение информации о пользователе
    user_id = message.from_user.id
    username = message.from_user.username
    first_name = message.from_user.first_name
    last_name = message.from_user.last_name
    # Проверка наличия данных о пользователе в базе данных на основе user_id
    user_data_exists = check_user_data_exists(user_id)
    if user_data_exists:
        # Данные о пользователе уже существуют, номер телефона не запрашиваем
        await message.answer(f"Спасибо, {contact.full_name}.\nВы уже зарегистрированы.")

        send_video_text = (
            "🙏 Благодарю тебя и желаю бодрого настроения!\n\n"
            "Приготовь коврик и не забывай, что прием пищи до тренировки за 1,5 часа! После можешь принять контрастный "
            "душ и намазать свое шикарное "
            "тело ароматным кремом 🧴.\n"
            "<a href='https://youtu.be/JXRep76T4yU'>📽 Видео 1</a>\n"
            "<a href='https://youtu.be/59npEilTIjg'>📽 Видео 2</a>\n\n"
            "Мой инстаграм для обратной связи @olya_mescheryakova\n\n"
            "Для возврата нажмите /start")
        await bot.send_message(message.chat.id, send_video_text, disable_web_page_preview=True)
    else:
        # Данные о пользователе отсутствуют, запрашиваем номер телефона и записываем данные
        number_phone = contact.phone_number
        date_now = datetime.datetime.now()
        # Запись данных о пользователе в базу данных
        recording_user_data(user_id, username, number_phone, first_name, last_name, date_now)
        send_video_text = (f"Спасибо, {contact.full_name}.За регистрацию\n"
                           "Приготовь коврик и не забывай, что прием пищи до тренировки за 1,5 часа! После можешь "
                           "принять контрастный душ и намазать свое шикарное тело ароматным кремом 🧴.\n"
                           "<a href='https://youtu.be/JXRep76T4yU'>Видео 1</a>\n"
                           "<a href='https://youtu.be/59npEilTIjg'>Видео 2</a>\n\n"
                           "Мой инстаграм для обратной связи @olya_mescheryakova\n\n"
                           "Для возврата нажмите /start")

        # Удаляем клавиатуру после получения номера
        await bot.send_message(message.chat.id, send_video_text, reply_markup=types.ReplyKeyboardRemove(),
                               disable_web_page_preview=True)


@dp.callback_query_handler(lambda callback_query: callback_query.data == "get_video")
async def get_video(callback_query: types.CallbackQuery):
    """
    Обработчик нажатия inline кнопки "get_video", который отправляет видео пользователям.
    Параметры:
    - callback_query (types. CallbackQuery): Объект колбэк-запроса, содержащий информацию о действии пользователя.
    Внутренние действия:
    1. Получает идентификатор пользователя из колбэк-запроса.
    2. Проверяет наличие данных о пользователе в базе данных на основе идентификатора пользователя.
    3. Если данные о пользователе существуют, отправляет видео и информацию о тренировке.
    4. Если данные о пользователе отсутствуют, предлагает пользователю пройти регистрацию.
    Пример использования:
    При нажатии пользователем inline кнопки "get_video", обработчик отправит видео или предложит
    пройти регистрацию, в зависимости от наличия данных о пользователе.
    """
    user_id = callback_query.from_user.id  # Получение информации о пользователе
    # Проверка наличия данных о пользователе в базе данных на основе user_id
    user_data_exists = check_user_data_exists(user_id)
    if user_data_exists:
        # Данные о пользователе уже существуют, сразу отправляем видео
        send_video_text = (
            "🙏 Благодарю тебя и желаю бодрого настроения!\n"
            "Приготовь коврик и не забывай, что прием пищи до тренировки за 1,5 часа! После можешь принять контрастный "
            "душ и намазать свое шикарное тело ароматным кремом 🧴.\n"
            "<a href='https://youtu.be/JXRep76T4yU'>Тренировка 1</a>\n"
            "<a href='https://youtu.be/59npEilTIjg'>Тренировка 2</a>\n\n"
            "Мой инстаграм для обратной связи @olya_mescheryakova\n\n"
            "Для возврата нажмите /start")
        await bot.send_message(callback_query.message.chat.id, send_video_text, disable_web_page_preview=True)
    else:
        text_message = ("📹 Для получения видео, необходимо пройти короткую регистрацию.\n\n"
                        "🔐 Регистрация проводится один раз.\n\n"
                        "<i>Для регистрации, пожалуйста, нажмите кнопку ниже, чтобы отправить контакт. 📱</i>")
        await bot.send_message(callback_query.message.chat.id, text_message, reply_markup=contact_keyboard())


@dp.message_handler(commands=['export_user'])
async def export_command(message: types.Message):
    """
    Обработчик команды export_user, который извлекает данные о пользователях из базы данных SQLite
    и экспортирует их в файл Excel (users.xlsx), который затем отправляется пользователю.
    Параметры:
    - message (types. Message): Объект сообщения пользователя, который вызвал команду.
    Внутренние действия:
    1. Устанавливается соединение с базой данных SQLite.
    2. Извлекаются данные всех пользователей из таблицы user_data в базе данных.
    3. Создается новый файл Excel (users.xlsx) и заголовки столбцов записываются в первую строку.
    4. Данные о пользователях записываются в соответствующие столбцы.
    5. Файл Excel сохраняется и отправляется пользователю в виде документа.
    6. Файл Excel удаляется после отправки.
    Пример использования:
    После отправки команды /export_user, данные о пользователях будут экспортированы
    в файл Excel и отправлены пользователю.
    """
    conn = sqlite3.connect(database)
    cursor = conn.cursor()
    # Получаем данные всех пользователей из базы данных
    cursor.execute('SELECT * FROM user_data')
    data = cursor.fetchall()
    # Создаем файл Excel и записываем данные
    wb = openpyxl.Workbook()
    sheet = wb.active
    # Записываем заголовки
    headers = ["user_id", "username", "number_phone", "first_name", "last_name", "date_now"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header
    # Записываем данные пользователей
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            sheet[f'{col_letter}{row_num}'] = cell_data
    wb.save('users.xlsx')  # Сохраняем файл Excel
    # Отправляем файл пользователю
    with open('users.xlsx', 'rb') as file:
        await bot.send_document(message.from_user.id, file, caption='Данные пользователей в формате Excel')
    os.remove('users.xlsx')  # Удаляем файл Excel


def greeting_handler():
    """Регистрируем handlers для бота (обработчиков сообщений)"""
    dp.register_message_handler(start_command)  # Обработчик команды /start, он же пост приветствия
    dp.register_message_handler(send_message)  # Обработчик команды /send_message, он же пост приветствия
    dp.register_message_handler(export_command)  # Обработчик команды /export_user, он же пост приветствия
